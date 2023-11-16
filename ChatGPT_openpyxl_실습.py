# 파이썬에서 openpyxl라이브러리를 사용해서 전자제품 판매데이터를 생성하는데 
# 전자제품ID, 이름, 가격, 수량을 sales.xlsx에 저장하는데 100개 샘플 데이터를 작성해줘. 
# 5살 아이가 이해할수 있도록 자세히 주석도 달아줘


import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Step 1: Create a new workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active

# Step 2: Add headers to the sheet (Electronic Product ID, Name, Price, Quantity)
headers = ['Electronic Product ID', 'Name', 'Price', 'Quantity']
for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    sheet[f'{col_letter}1'] = header

# Step 3: Generate 100 sample data entries
for row_num in range(2, 102):  # Rows 2 to 101 (100 entries)
    # Electronic Product ID: It can be a simple incremental number for this example
    sheet[f'A{row_num}'] = row_num - 1

    # Name: Assume electronic products have names like Product1, Product2, etc.
    sheet[f'B{row_num}'] = f'Product{row_num - 1}'

    # Price: Random price between $10 and $100 for simplicity
    price = round((10 + row_num - 2) * 1.5, 2)
    sheet[f'C{row_num}'] = price

    # Quantity: Random quantity between 1 and 10 for simplicity
    quantity = row_num % 10 + 1
    sheet[f'D{row_num}'] = quantity

# Step 4: Save the workbook to a file named 'sales.xlsx'
workbook.save('sales.xlsx')

# Step 5: Print a message indicating that the file has been created

